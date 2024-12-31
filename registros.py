import tkinter as tk
from openpyxl import Workbook, load_workbook
import tkinter.messagebox as messagebox
from tkcalendar import DateEntry

class RegistroBancario:
    def __init__(self):
        self.notebook = tk.Tk()
        self.notebook.title('Aplicación de Registros Bancarios')
        self.ventana_agregar_registro = tk.Frame(self.notebook)
        self.ventana_visualizar_registro = None

        # Crear un menú
        self.menu_bar = tk.Menu(self.notebook)

        # Agregar opciones al menú
        self.opcion_agregar_registro = tk.Menu(self.menu_bar, tearoff=0)
        self.opcion_agregar_registro.add_command(label='Agregar Registro', command=self.agregar_registro)
        self.opcion_agregar_registro.add_separator()
        self.opcion_agregar_registro.add_command(label='Visualizar Registros', command=self.visualizar_registros)

        # Agregar opción para salir
        self.opcion_salir = tk.Menu(self.menu_bar, tearoff=0)
        self.opcion_salir.add_command(label='Salir', command=self.notebook.destroy)

        # Agregar opciones al menú principal
        self.menu_bar.add_cascade(label='Archivo', menu=self.opcion_agregar_registro)
        self.menu_bar.add_cascade(label='Salir', menu=self.opcion_salir)

        # Agregar el menú a la ventana principal
        self.notebook.config(menu=self.menu_bar)

    def agregar_registro(self):
        self.ventana_agregar_registro = tk.Frame(self.notebook)
        self.ventana_agregar_registro.pack(fill='both', expand=True)

        # Crear etiquetas y entradas para los valores del registro
        self.label_lugar = tk.Label(self.ventana_agregar_registro, text="LUGAR")
        self.label_lugar.grid(row=1, column=0)
        self.entrada_lugar = tk.Entry(self.ventana_agregar_registro, width=50)
        self.entrada_lugar.grid(row=1, column=1)

        self.label_fecha = tk.Label(self.ventana_agregar_registro, text="FECHA")
        self.label_fecha.grid(row=2, column=0)
        self.calendario_fecha = DateEntry(self.ventana_agregar_registro, width=12, background='darkblue', foreground='white', borderwidth=0)
        self.calendario_fecha.grid(row=2, column=1)

        self.label_cantidad = tk.Label(self.ventana_agregar_registro, text="CANTIDAD")
        self.label_cantidad.grid(row=3, column=0)
        self.entrada_cantidad = tk.Entry(self.ventana_agregar_registro, width=50)
        self.entrada_cantidad.grid(row=3, column=1)

        self.label_precio_original = tk.Label(self.ventana_agregar_registro, text="PRECIO ORIGINAL")
        self.label_precio_original.grid(row=4, column=0)
        self.entrada_precio_original = tk.Entry(self.ventana_agregar_registro, width=50)
        self.entrada_precio_original.grid(row=4, column=1)

        self.label_precio_total = tk.Label(self.ventana_agregar_registro, text="PRECIO TOTAL")
        self.label_precio_total.grid(row=5, column=0)
        self.entrada_precio_total = tk.Entry(self.ventana_agregar_registro, width=50)
        self.entrada_precio_total.grid(row=5, column=1)

        self.label_precio_cuota = tk.Label(self.ventana_agregar_registro, text="PRECIO CUOTA")
        self.label_precio_cuota.grid(row=6, column=0)
        self.entrada_precio_cuota = tk.Entry(self.ventana_agregar_registro, width=50)
        self.entrada_precio_cuota.grid(row=6, column=1)

        self.label_cuotas_restantes = tk.Label(self.ventana_agregar_registro, text="CUOTAS RESTANTES")
        self.label_cuotas_restantes.grid(row=7, column=0)
        self.entrada_cuotas_restantes = tk.Entry(self.ventana_agregar_registro, width=50)
        self.entrada_cuotas_restantes.grid(row=7, column=1)

        self.label_ticket = tk.Label(self.ventana_agregar_registro, text="TICKET")
        self.label_ticket.grid(row=8, column=0)
        self.entrada_ticket = tk.Entry(self.ventana_agregar_registro, width=50)
        self.entrada_ticket.grid(row=8, column=1)

        self.label_tarjeta = tk.Label(self.ventana_agregar_registro, text="TARJETA")
        self.label_tarjeta.grid(row=9, column=0)
        self.entrada_tarjeta = tk.Entry(self.ventana_agregar_registro, width=50)
        self.entrada_tarjeta.grid(row=9, column=1)

        self.label_cuotas = tk.Label(self.ventana_agregar_registro, text="CUOTAS")
        self.label_cuotas.grid(row=10, column=0)
        self.entrada_cuotas = tk.Entry(self.ventana_agregar_registro, width=50)
        self.entrada_cuotas.grid(row=10, column=1)

        # Botón para agregar registro
        self.boton_agregar_registro = tk.Button(self.ventana_agregar_registro, text="Agregar Registro", command=self.agregar_registro_aceptar)
        self.boton_agregar_registro.grid(row=11, column=0, columnspan=2, padx=10, pady=10)

    def agregar_registro_aceptar(self):
        lugar = self.entrada_lugar.get()
        fecha = self.calendario_fecha.get_date().strftime("%d/%m/%Y")
        cantidad = self.entrada_cantidad.get()
        precio_original = self.entrada_precio_original.get()
        precio_total = self.entrada_precio_total.get()
        precio_cuota = self.entrada_precio_cuota.get()
        cuotas_restantes = self.entrada_cuotas_restantes.get()
        ticket = self.entrada_ticket.get()
        tarjeta = self.entrada_tarjeta.get()
        cuotas = self.entrada_cuotas.get()

        if not lugar or not fecha or not cantidad or not precio_original or not precio_total or not precio_cuota or not cuotas_restantes or not ticket or not tarjeta or not cuotas:
            messagebox.showerror("Error", "Por favor, complete todos los campos")
            return

        try:
            cantidad = int(cantidad)
            precio_original = float(precio_original)
            precio_total = float(precio_total)
            precio_cuota = float(precio_cuota)
            cuotas_restantes = int(cuotas_restantes)
        except ValueError:
            messagebox.showerror("Error", "Los campos de cantidad, precio original, precio total y cuotas restantes deben ser números")
            return

        if cantidad <= 0 or precio_original <= 0 or precio_total <= 0 or precio_cuota <= 0 or cuotas_restantes <= 0:
            messagebox.showerror("Error", "Los campos de cantidad, precio original, precio total y cuotas restantes deben ser números positivos")
            return

        # Abrir el archivo de Excel
        try:
            self.archivo_excel = load_workbook(filename='registrosbancarios.xlsx')
        except FileNotFoundError:
            self.archivo_excel = Workbook()
        
        # Agregar registro al libro de Excel
        sheet = self.archivo_excel.active
        sheet.append([lugar, fecha, cantidad, precio_original, precio_total, precio_cuota, cuotas_restantes, ticket, tarjeta, cuotas])

        # Guardar el archivo de Excel
        self.archivo_excel.save('registrosbancarios.xlsx')

    def visualizar_registros(self):
        try:
            self.archivo_excel = load_workbook(filename='registrosbancarios.xlsx')
        except FileNotFoundError:
            messagebox.showerror("Error", "No hay registros para mostrar")
            return
        
        # Crear un cuadro de texto para mostrar los registros
        self.ventana_visualizar_registro = tk.Frame(self.notebook)
        self.ventana_visualizar_registro.pack(fill='both', expand=True)

        # Mostrar los registros en el cuadro de texto
        sheet = self.archivo_excel.active
        self.entrada_registros = tk.Text(self.ventana_visualizar_registro, width=100, height=20)
        self.entrada_registros.pack(fill='both', expand=True)
        
        for i, row in enumerate(sheet.rows):
            registro = ', '.join(str(cell.value) for cell in row)
            if i > 0:
                self.entrada_registros.insert(tk.END, '\n')
            self.entrada_registros.insert(tk.END, registro + '\n')

    def run(self):
        self.notebook.mainloop()

if __name__ == '__main__':
    app = RegistroBancario()
    app.run()

